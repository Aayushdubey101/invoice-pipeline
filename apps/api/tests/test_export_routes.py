"""
Excel export regression tests:
  - confidence_breakdown key must be "overall_score" (not "overall_confidence") or
    the Confidence column silently exports blank for every row.
  - business columns (vendor tax id/address, buyer, payment terms, PO number) must
    be present since the extractor already captures them but export used to drop them.
"""
from decimal import Decimal

import openpyxl
import pytest
from httpx import ASGITransport, AsyncClient
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from datetime import datetime, timedelta, timezone

from invoice_pipeline.api.main import app
from invoice_pipeline.db.models import Base, Document, Invoice, Vendor, Workspace
from invoice_pipeline.db.session import get_session


@pytest.fixture
async def db_engine():
    engine = create_async_engine("sqlite+aiosqlite:///:memory:", echo=False)
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    yield engine
    await engine.dispose()


@pytest.fixture
async def db_session(db_engine):
    factory = async_sessionmaker(db_engine, expire_on_commit=False)
    async with factory() as session:
        yield session


@pytest.fixture
async def workspace(db_session: AsyncSession) -> Workspace:
    w = Workspace(workspace_type="guest", expires_at=datetime.now(timezone.utc) + timedelta(hours=1))
    db_session.add(w)
    await db_session.commit()
    return w


@pytest.fixture
async def client(db_session: AsyncSession, workspace: Workspace):
    async def override_get_session():
        yield db_session

    app.dependency_overrides[get_session] = override_get_session
    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://test",
        headers={"X-Workspace-Id": workspace.id},
    ) as ac:
        yield ac
    app.dependency_overrides.clear()


@pytest.mark.asyncio
async def test_export_excel_includes_business_fields_and_confidence(
    client: AsyncClient, db_session: AsyncSession, workspace: Workspace
) -> None:
    vendor = Vendor(
        id="v-001",
        workspace_id=workspace.id,
        canonical_name="Acme Corp",
        tax_id="TAX-999",
        address="1 Acme Way",
    )
    doc = Document(
        id="d" * 64,
        workspace_id=workspace.id,
        filename="invoice.pdf",
        mime_type="application/pdf",
        file_size_bytes=1024,
        doc_type="text_pdf",
        status="complete",
        errors=[],
    )
    db_session.add_all([vendor, doc])
    await db_session.flush()

    inv = Invoice(
        id="inv-001",
        workspace_id=workspace.id,
        document_id=doc.id,
        vendor_id=vendor.id,
        invoice_number="INV-001",
        invoice_date="2024-01-15",
        buyer_name="Beta Inc",
        payment_terms="Net 30",
        purchase_order="PO-42",
        subtotal=Decimal("1000.00"),
        tax_amount=Decimal("100.00"),
        total_amount=Decimal("1100.00"),
        currency="USD",
        needs_review=False,
        review_reasons=[],
        raw_extraction={},
        confidence_breakdown={"overall_score": 0.87},
    )
    db_session.add(inv)
    await db_session.commit()

    res = await client.get("/export/excel")
    assert res.status_code == 200

    import io

    wb = openpyxl.load_workbook(io.BytesIO(res.content))
    ws = wb["Invoices"]
    headers = [c.value for c in ws[1]]

    for expected in ("Vendor Tax ID", "Vendor Address", "Buyer", "Payment Terms", "PO Number"):
        assert expected in headers, f"missing column: {expected}"

    row = {headers[i]: v for i, v in enumerate(next(ws.iter_rows(min_row=2, max_row=2, values_only=True)))}
    assert row["Vendor Tax ID"] == "TAX-999"
    assert row["Vendor Address"] == "1 Acme Way"
    assert row["Buyer"] == "Beta Inc"
    assert row["Payment Terms"] == "Net 30"
    assert row["PO Number"] == "PO-42"
    assert row["Confidence"] == pytest.approx(0.87)
