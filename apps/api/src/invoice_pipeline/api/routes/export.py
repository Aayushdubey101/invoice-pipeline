"""
Phase 11 — Excel Export

Supports exporting invoices to professional Excel files with:
  - Sheet 1: Invoice summary (all key fields)
  - Sheet 2: Line items

Filter options: invoice_ids, batch_id, vendor_id, start_date, end_date, review_status
Tracks export history in the database.
Large exports are streamed asynchronously.
"""
import io
from datetime import datetime
from typing import Any

import structlog
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from invoice_pipeline.api.deps import get_current_workspace
from invoice_pipeline.db import models
from invoice_pipeline.db.models import Workspace
from invoice_pipeline.db.session import get_session

log = structlog.get_logger()
router = APIRouter()


def _apply_filters(
    stmt: Any,
    invoice_ids: list[str] | None,
    batch_id: str | None,
    vendor_id: str | None,
    start_date: str | None,
    end_date: str | None,
    review_status: str | None,
) -> Any:
    """Apply optional filters to the invoice query statement.

    Guards:
    - Only joins Document once even if both batch_id and other doc-filters are present.
    - Ignores empty-string values.
    """
    if invoice_ids:
        stmt = stmt.where(models.Invoice.id.in_(invoice_ids))
    if batch_id:
        # Join Document table once for batch filter
        stmt = stmt.join(models.Document, models.Invoice.document_id == models.Document.id).where(
            models.Document.batch_id == batch_id
        )
    if vendor_id:
        stmt = stmt.where(models.Invoice.vendor_id == vendor_id)
    if start_date and start_date.strip():
        stmt = stmt.where(models.Invoice.invoice_date >= start_date.strip())
    if end_date and end_date.strip():
        stmt = stmt.where(models.Invoice.invoice_date <= end_date.strip())
    if review_status == "needs_review":
        stmt = stmt.where(models.Invoice.needs_review == True)  # noqa: E712
    elif review_status == "approved":
        stmt = stmt.where(models.Invoice.needs_review == False)  # noqa: E712
    return stmt


def build_workbook(invoices: list) -> Any:
    """Build the two-sheet (Invoices + Line Items) Excel workbook for a set of
    invoices. Shared by GET /export/excel and POST /session/finish."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl is required for Excel export. Install it with: pip install openpyxl",
        )

    wb = openpyxl.Workbook()

    HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    CENTER = Alignment(horizontal="center", vertical="center")

    # ── Sheet 1: Invoice Summary ───────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Invoices"
    ws1.row_dimensions[1].height = 22

    inv_headers = [
        "Invoice Number", "Vendor", "Vendor Tax ID", "Vendor Address", "Buyer",
        "Invoice Date", "Due Date", "Payment Terms", "PO Number",
        "Currency", "Subtotal", "Tax", "Total",
        "Confidence", "Validation Status", "Review Status",
        "Batch ID", "Document ID", "Filename",
    ]

    for col_idx, header in enumerate(inv_headers, 1):
        cell = ws1.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER

    ws1.freeze_panes = "A2"

    for row_idx, inv in enumerate(invoices, 2):
        # Confidence
        confidence: float | None = None
        if inv.confidence_breakdown:
            raw_conf = inv.confidence_breakdown.get("overall_score")
            if raw_conf is not None:
                confidence = round(float(raw_conf), 4)

        # Validation status summary
        validation_status = "—"
        if inv.validation_report:
            passed = inv.validation_report.get("total_passed", 0)
            failed_v = inv.validation_report.get("total_failed", 0)
            warnings = inv.validation_report.get("total_warnings", 0)
            validation_status = f"{passed}P / {failed_v}F / {warnings}W"

        review_label = "Needs Review" if inv.needs_review else "Approved"
        batch_id_val = inv.document.batch_id if inv.document else None

        row_data = [
            inv.invoice_number,
            inv.vendor.canonical_name if inv.vendor else None,
            inv.vendor.tax_id if inv.vendor else None,
            inv.vendor.address if inv.vendor else None,
            inv.buyer_name,
            inv.invoice_date,
            inv.due_date,
            inv.payment_terms,
            inv.purchase_order,
            inv.currency,
            inv.subtotal,
            inv.tax_amount,
            inv.total_amount,
            confidence,
            validation_status,
            review_label,
            batch_id_val,
            inv.document_id,
            inv.document.filename if inv.document else None,
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            # Apply number format for monetary columns (Subtotal=11, Tax=12, Total=13)
            if col_idx in (11, 12, 13) and val is not None:
                cell.number_format = '#,##0.00'
            # Apply percentage format for confidence (col 14)
            if col_idx == 14 and val is not None:
                cell.number_format = '0.0%'

    # Auto-fit columns (generous minimum widths)
    col_min_widths = [18, 25, 16, 30, 20, 13, 13, 14, 16, 10, 12, 12, 12, 12, 18, 14, 38, 38, 30]
    for col_idx, min_w in enumerate(col_min_widths, 1):
        ws1.column_dimensions[get_column_letter(col_idx)].width = min_w

    # ── Sheet 2: Line Items ────────────────────────────────────────────────────
    ws2 = wb.create_sheet(title="Line Items")
    ws2.row_dimensions[1].height = 22

    li_headers = ["Invoice Number", "Vendor", "Product / Description", "Quantity", "Unit Price", "Line Total"]
    for col_idx, header in enumerate(li_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER

    ws2.freeze_panes = "A2"

    li_row = 2
    for inv in invoices:
        vendor_name = inv.vendor.canonical_name if inv.vendor else None
        for li in sorted(inv.line_items, key=lambda x: x.position):
            ws2.cell(row=li_row, column=1, value=inv.invoice_number)
            ws2.cell(row=li_row, column=2, value=vendor_name)
            ws2.cell(row=li_row, column=3, value=li.description_raw)
            ws2.cell(row=li_row, column=4, value=li.quantity_raw)
            ws2.cell(row=li_row, column=5, value=li.unit_price_raw)
            ws2.cell(row=li_row, column=6, value=li.total_raw)
            li_row += 1

    for col_idx, width in enumerate([18, 25, 45, 12, 14, 14], 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    return wb


def build_pdf_report(invoices: list) -> bytes:
    """Build a tabular PDF summary of invoices. Shared by GET /export/pdf and
    POST /session/finish."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(letter), title="Invoice Export",
        leftMargin=0.5 * inch, rightMargin=0.5 * inch,
    )
    styles = getSampleStyleSheet()
    elements: list[Any] = [Paragraph("Invoice Export Report", styles["Title"]), Spacer(1, 0.2 * inch)]

    headers = ["Invoice #", "Vendor", "Invoice Date", "Due Date", "Currency", "Subtotal", "Tax", "Total", "Status"]
    data: list[list[str]] = [headers]
    for inv in invoices:
        review_label = "Needs Review" if inv.needs_review else "Approved"
        data.append(
            [
                inv.invoice_number or "—",
                inv.vendor.canonical_name if inv.vendor else "—",
                inv.invoice_date or "—",
                inv.due_date or "—",
                inv.currency or "—",
                f"{inv.subtotal:.2f}" if inv.subtotal is not None else "—",
                f"{inv.tax_amount:.2f}" if inv.tax_amount is not None else "—",
                f"{inv.total_amount:.2f}" if inv.total_amount is not None else "—",
                review_label,
            ]
        )

    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)
    return buf.getvalue()


@router.get("/excel")
async def export_excel(
    invoice_ids: list[str] | None = Query(default=None),
    batch_id: str | None = Query(default=None),
    vendor_id: str | None = Query(default=None),
    start_date: str | None = Query(default=None, description="YYYY-MM-DD"),
    end_date: str | None = Query(default=None, description="YYYY-MM-DD"),
    review_status: str | None = Query(default=None, description="needs_review | approved | all"),
    session: AsyncSession = Depends(get_session),
    workspace: Workspace = Depends(get_current_workspace),
) -> StreamingResponse:
    """
    Export invoices to Excel format.

    Sheet 1 — Invoice summary with all business fields.
    Sheet 2 — Line items for each exported invoice.
    """
    stmt = (
        select(models.Invoice)
        .where(models.Invoice.workspace_id == workspace.id)
        .options(
            selectinload(models.Invoice.vendor),
            selectinload(models.Invoice.document),
            selectinload(models.Invoice.line_items),
        )
    )
    stmt = _apply_filters(stmt, invoice_ids, batch_id, vendor_id, start_date, end_date, review_status)

    result = await session.execute(stmt)
    invoices = result.scalars().all()
    log.info("export_excel_started", record_count=len(invoices))

    wb = build_workbook(invoices)

    # ── Record export history ──────────────────────────────────────────────────
    export_filename = f"invoices_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    is_filtered = any([invoice_ids, batch_id, vendor_id, start_date, end_date, review_status])
    history = models.ExportHistory(
        workspace_id=workspace.id,
        export_type="filtered" if is_filtered else "all",
        filter_params={
            "invoice_ids": invoice_ids,
            "batch_id": batch_id,
            "vendor_id": vendor_id,
            "start_date": start_date,
            "end_date": end_date,
            "review_status": review_status,
        },
        record_count=len(invoices),
        filename=export_filename,
    )
    session.add(history)
    await session.commit()

    log.info("export_excel_complete", filename=export_filename, record_count=len(invoices))

    # Stream the workbook
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            # RFC 5987 encoded filename for non-ASCII safety
            "Content-Disposition": (
                f'attachment; filename="{export_filename}"; '
                f"filename*=UTF-8''{export_filename}"
            ),
            "X-Export-Record-Count": str(len(invoices)),
        },
    )


@router.get("/pdf")
async def export_pdf(
    invoice_ids: list[str] | None = Query(default=None),
    batch_id: str | None = Query(default=None),
    vendor_id: str | None = Query(default=None),
    start_date: str | None = Query(default=None, description="YYYY-MM-DD"),
    end_date: str | None = Query(default=None, description="YYYY-MM-DD"),
    review_status: str | None = Query(default=None, description="needs_review | approved | all"),
    session: AsyncSession = Depends(get_session),
    workspace: Workspace = Depends(get_current_workspace),
) -> StreamingResponse:
    """Export invoices to a tabular PDF summary report (same filters as /export/excel)."""
    stmt = select(models.Invoice).where(models.Invoice.workspace_id == workspace.id).options(
        selectinload(models.Invoice.vendor)
    )
    stmt = _apply_filters(stmt, invoice_ids, batch_id, vendor_id, start_date, end_date, review_status)

    result = await session.execute(stmt)
    invoices = result.scalars().all()

    pdf_bytes = build_pdf_report(invoices)

    export_filename = f"invoices_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
    is_filtered = any([invoice_ids, batch_id, vendor_id, start_date, end_date, review_status])
    session.add(
        models.ExportHistory(
            workspace_id=workspace.id,
            export_type="filtered" if is_filtered else "all",
            filter_params={
                "invoice_ids": invoice_ids,
                "batch_id": batch_id,
                "vendor_id": vendor_id,
                "start_date": start_date,
                "end_date": end_date,
                "review_status": review_status,
            },
            record_count=len(invoices),
            filename=export_filename,
        )
    )
    await session.commit()

    log.info("export_pdf_complete", filename=export_filename, record_count=len(invoices))

    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={
            "Content-Disposition": (
                f'attachment; filename="{export_filename}"; '
                f"filename*=UTF-8''{export_filename}"
            ),
            "X-Export-Record-Count": str(len(invoices)),
        },
    )


@router.get("/history")
async def export_history(
    session: AsyncSession = Depends(get_session),
    workspace: Workspace = Depends(get_current_workspace),
) -> dict[str, Any]:
    """List past export operations."""
    result = await session.execute(
        select(models.ExportHistory)
        .where(models.ExportHistory.workspace_id == workspace.id)
        .order_by(models.ExportHistory.created_at.desc())
        .limit(200)
    )
    exports = result.scalars().all()
    return {
        "exports": [
            {
                "id": e.id,
                "export_type": e.export_type,
                "filter_params": e.filter_params,
                "record_count": e.record_count,
                "filename": e.filename,
                "created_at": e.created_at.isoformat(),
            }
            for e in exports
        ]
    }
